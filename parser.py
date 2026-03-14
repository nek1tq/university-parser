"""
US News Best Global Universities Parser
Парсит 2551 университет с usnews.com и сохраняет статистику в Excel.

Шаг 1: Получить список всех университетов через API (search endpoint)
Шаг 2: Для каждого университета загрузить страницу и извлечь статистику из HTML
Шаг 3: Сохранить в Excel по шаблону заказчика
"""

import requests
import re
import time
import json
import html
import openpyxl
from pathlib import Path

# --- Настройки ---
SEARCH_API = "https://www.usnews.com/education/best-global-universities/api/search"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.usnews.com/",
}
OUTPUT_FILE = "US_News_2025-2026.xlsx"
PROGRESS_FILE = "progress.json"
DELAY = 1.0  # секунд между запросами

# Колонки из шаблона заказчика
COLUMNS = [
    "University Name",
    "Total number of students",
    "Link",
    "Number of international students",
    "Total number of academic staff",
    "Number of international staff",
    "Number of undergraduate degrees awarded",
    "Number of master's degrees awarded",
    "Number of doctoral degrees awarded",
    "Number of research only staff",
    "Number of new undergraduate students",
    "Number of new master's students",
    "Number of new doctoral students",
]

# Метки которые ищем на странице -> колонка в Excel
STAT_LABELS = {
    "Total number of students": "Total number of students",
    "Number of international students": "Number of international students",
    "Total number of academic staff": "Total number of academic staff",
    "Number of international staff": "Number of international staff",
    "Number of undergraduate degrees awarded": "Number of undergraduate degrees awarded",
    "Number of master's degrees awarded": "Number of master's degrees awarded",
    "Number of master&#x27;s degrees awarded": "Number of master's degrees awarded",
    "Number of doctoral degrees awarded": "Number of doctoral degrees awarded",
    "Number of research only staff": "Number of research only staff",
    "Number of new undergraduate students": "Number of new undergraduate students",
    "Number of new master's students": "Number of new master's students",
    "Number of new master&#x27;s students": "Number of new master's students",
    "Number of new doctoral students": "Number of new doctoral students",
}


def get_all_universities():
    """Шаг 1: Получить список всех университетов через API."""
    all_unis = []
    page = 1

    while True:
        print(f"  Загружаю страницу {page}...", end=" ")
        try:
            r = requests.get(
                SEARCH_API,
                params={"format": "json", "page": page},
                headers=HEADERS,
                timeout=30,
            )
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"Ошибка: {e}")
            time.sleep(5)
            continue

        items = data.get("items", [])
        if not items:
            break

        for item in items:
            all_unis.append({
                "name": item["name"],
                "url": item["url"],
            })

        print(f"OK ({len(items)} университетов, всего {len(all_unis)})")

        total_pages = data.get("total_pages", 0)
        if page >= total_pages:
            break

        page += 1
        time.sleep(0.5)

    return all_unis


def parse_university_page(url):
    """Шаг 2: Загрузить страницу университета и извлечь статистику."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        return {"error": str(e)}

    text = r.text
    stats = {}

    # Паттерн: >Label</p><p class="...">Value</p>
    pattern = r">([^<]+)</p><p[^>]*>([^<]+)</p>"
    matches = re.findall(pattern, text)

    for raw_label, value in matches:
        label = html.unescape(raw_label).strip()
        if label in STAT_LABELS:
            col_name = STAT_LABELS[label]
            # Убираем запятые из чисел
            clean_value = value.strip().replace(",", "")
            try:
                stats[col_name] = int(clean_value)
            except ValueError:
                stats[col_name] = value.strip()

    return stats


def load_progress():
    """Загрузить прогресс из файла (для возобновления после обрыва)."""
    if Path(PROGRESS_FILE).exists():
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"completed": {}, "universities": []}


def save_progress(progress):
    """Сохранить прогресс."""
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def save_to_excel(progress):
    """Шаг 3: Сохранить результаты в Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Universities"

    # Заголовки
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Данные
    row_idx = 2
    for uni in progress["universities"]:
        url = uni["url"]
        stats = progress["completed"].get(url, {})

        ws.cell(row=row_idx, column=1, value=uni["name"])
        ws.cell(row=row_idx, column=2, value=stats.get("Total number of students", ""))
        ws.cell(row=row_idx, column=3, value=url)
        ws.cell(row=row_idx, column=4, value=stats.get("Number of international students", ""))
        ws.cell(row=row_idx, column=5, value=stats.get("Total number of academic staff", ""))
        ws.cell(row=row_idx, column=6, value=stats.get("Number of international staff", ""))
        ws.cell(row=row_idx, column=7, value=stats.get("Number of undergraduate degrees awarded", ""))
        ws.cell(row=row_idx, column=8, value=stats.get("Number of master's degrees awarded", ""))
        ws.cell(row=row_idx, column=9, value=stats.get("Number of doctoral degrees awarded", ""))
        ws.cell(row=row_idx, column=10, value=stats.get("Number of research only staff", ""))
        ws.cell(row=row_idx, column=11, value=stats.get("Number of new undergraduate students", ""))
        ws.cell(row=row_idx, column=12, value=stats.get("Number of new master's students", ""))
        ws.cell(row=row_idx, column=13, value=stats.get("Number of new doctoral students", ""))

        row_idx += 1

    wb.save(OUTPUT_FILE)
    print(f"\nСохранено в {OUTPUT_FILE} ({row_idx - 2} строк)")


def main():
    print("=" * 60)
    print("US News — Парсер университетов")
    print("=" * 60)

    # Загрузить прогресс
    progress = load_progress()

    # Шаг 1: Получить список
    if not progress["universities"]:
        print("\n[1/3] Получаю список университетов...")
        unis = get_all_universities()
        progress["universities"] = unis
        save_progress(progress)
        print(f"Найдено {len(unis)} университетов")
    else:
        unis = progress["universities"]
        print(f"\n[1/3] Список загружен: {len(unis)} университетов")

    # Шаг 2: Парсить каждый
    completed = len(progress["completed"])
    total = len(unis)
    print(f"\n[2/3] Парсинг страниц ({completed}/{total} уже готово)...")

    for i, uni in enumerate(unis):
        url = uni["url"]

        if url in progress["completed"]:
            continue

        print(f"  [{i+1}/{total}] {uni['name'][:50]}...", end=" ")

        stats = parse_university_page(url)

        if "error" in stats:
            print(f"ОШИБКА: {stats['error']}")
            # Retry once after delay
            time.sleep(5)
            stats = parse_university_page(url)
            if "error" in stats:
                print(f"  Повторная ошибка, пропускаю")
                stats = {}

        progress["completed"][url] = stats
        found = len([v for v in stats.values() if v])
        print(f"OK ({found}/11 метрик)")

        # Сохраняем прогресс каждые 10 университетов
        if (i + 1) % 10 == 0:
            save_progress(progress)
            save_to_excel(progress)
            print(f"  --- Прогресс сохранён: {len(progress['completed'])}/{total} ---")

        time.sleep(DELAY)

    # Финальное сохранение
    save_progress(progress)

    # Шаг 3: Excel
    print(f"\n[3/3] Сохраняю в Excel...")
    save_to_excel(progress)

    print("\nГотово!")


if __name__ == "__main__":
    main()
