"""
Автотесты для university-parser.
Учимся: pytest, моки (unittest.mock), фикстуры, параметризация.

Запуск: pytest test_parser.py -v
"""

import json
import pytest
from unittest.mock import patch, MagicMock
from pathlib import Path

# Импортируем функции из нашего парсера
from parser import (
    parse_university_page,
    load_progress,
    save_progress,
    save_to_excel,
    COLUMNS,
    STAT_LABELS,
)


# ============================================================
# ФИКСТУРЫ — подготовка данных для тестов
# ============================================================
# Фикстура — это функция которая создаёт тестовые данные.
# pytest автоматически вызывает её перед каждым тестом
# который принимает аргумент с таким же именем.

@pytest.fixture
def sample_progress():
    """Создаёт пример прогресса с двумя университетами."""
    return {
        "completed": {
            "https://www.usnews.com/education/best-global-universities/harvard-university": {
                "Total number of students": 36012,
                "Number of international students": 7744,
                "Total number of academic staff": 4184,
            },
            "https://www.usnews.com/education/best-global-universities/mit": {},
        },
        "universities": [
            {"name": "Harvard University", "url": "https://www.usnews.com/education/best-global-universities/harvard-university"},
            {"name": "Massachusetts Institute of Technology", "url": "https://www.usnews.com/education/best-global-universities/mit"},
        ],
    }


@pytest.fixture
def sample_html():
    """HTML-фрагмент как на реальной странице usnews.com."""
    return """
    <div class="DataRow">
        <p class="label">Total number of students</p><p class="value">36,012</p>
    </div>
    <div class="DataRow">
        <p class="label">Number of international students</p><p class="value">7,744</p>
    </div>
    <div class="DataRow">
        <p class="label">Number of master&#x27;s degrees awarded</p><p class="value">5,501</p>
    </div>
    <div class="DataRow">
        <p class="label">Number of doctoral degrees awarded</p><p class="value">1,234</p>
    </div>
    """


@pytest.fixture
def tmp_progress_file(tmp_path):
    """
    tmp_path — встроенная фикстура pytest.
    Создаёт временную папку которая удаляется после теста.
    Так мы не мусорим в реальной файловой системе.
    """
    return tmp_path / "test_progress.json"


# ============================================================
# ТЕСТ 1: Парсинг HTML страницы
# ============================================================
# Почему это важно: это ядро парсера. Если parse_university_page
# сломается — весь парсер бесполезен.

class TestParseUniversityPage:
    """Группа тестов для parse_university_page."""

    def test_parses_numbers_correctly(self, sample_html):
        """
        Проверяем что парсер:
        1) Находит метки в HTML
        2) Убирает запятые из чисел (36,012 → 36012)
        3) Конвертирует в int
        """
        # patch — подменяем requests.get чтобы не делать реальный запрос
        # Это называется "мок" (mock) — имитация внешней зависимости
        with patch("parser.requests.get") as mock_get:
            # Настраиваем мок: что он вернёт
            mock_response = MagicMock()
            mock_response.text = sample_html
            mock_response.raise_for_status = MagicMock()  # не кидает ошибку
            mock_get.return_value = mock_response

            # Вызываем функцию
            result = parse_university_page("https://fake-url.com")

            # Проверяем результат
            assert result["Total number of students"] == 36012
            assert result["Number of international students"] == 7744
            assert result["Number of doctoral degrees awarded"] == 1234

    def test_handles_html_entities(self, sample_html):
        """
        Проверяем что master&#x27;s декодируется в master's.
        Это реальный баг который мы ловили при разработке!
        """
        with patch("parser.requests.get") as mock_get:
            mock_response = MagicMock()
            mock_response.text = sample_html
            mock_response.raise_for_status = MagicMock()
            mock_get.return_value = mock_response

            result = parse_university_page("https://fake-url.com")

            # Ключ должен быть с апострофом, не с HTML-entity
            assert "Number of master's degrees awarded" in result
            assert result["Number of master's degrees awarded"] == 5501

    def test_handles_network_error(self):
        """
        Проверяем что при ошибке сети функция не падает,
        а возвращает словарь с ключом 'error'.
        """
        with patch("parser.requests.get") as mock_get:
            # Имитируем ошибку сети
            mock_get.side_effect = Exception("Connection timeout")

            result = parse_university_page("https://fake-url.com")

            assert "error" in result
            assert "Connection timeout" in result["error"]

    def test_empty_page_returns_empty_dict(self):
        """Страница без данных → пустой словарь (не ошибка)."""
        with patch("parser.requests.get") as mock_get:
            mock_response = MagicMock()
            mock_response.text = "<html><body>No data here</body></html>"
            mock_response.raise_for_status = MagicMock()
            mock_get.return_value = mock_response

            result = parse_university_page("https://fake-url.com")

            assert result == {}
            assert "error" not in result


# ============================================================
# ТЕСТ 2: Сохранение и загрузка прогресса
# ============================================================
# Почему это важно: если прогресс сломается — потеряем часы работы
# парсера и придётся начинать заново.

class TestProgress:

    def test_save_and_load(self, tmp_progress_file, sample_progress):
        """Сохраняем прогресс → загружаем → данные совпадают."""
        # Подменяем путь к файлу прогресса
        with patch("parser.PROGRESS_FILE", str(tmp_progress_file)):
            save_progress(sample_progress)

            # Проверяем что файл создан
            assert tmp_progress_file.exists()

            # Загружаем и сравниваем
            loaded = load_progress()
            assert loaded == sample_progress

    def test_load_missing_file_returns_default(self, tmp_path):
        """Если файла нет — возвращаем пустую структуру, не падаем."""
        fake_path = str(tmp_path / "nonexistent.json")
        with patch("parser.PROGRESS_FILE", fake_path):
            result = load_progress()

            assert result == {"completed": {}, "universities": []}

    def test_progress_file_is_valid_json(self, tmp_progress_file, sample_progress):
        """Проверяем что файл — валидный JSON (не битый)."""
        with patch("parser.PROGRESS_FILE", str(tmp_progress_file)):
            save_progress(sample_progress)

            # Читаем как обычный текст и парсим JSON
            raw_text = tmp_progress_file.read_text(encoding="utf-8")
            parsed = json.loads(raw_text)  # Упадёт если JSON битый

            assert "completed" in parsed
            assert "universities" in parsed


# ============================================================
# ТЕСТ 3: Экспорт в Excel
# ============================================================

class TestExcelExport:

    def test_creates_file(self, tmp_path, sample_progress):
        """Excel файл создаётся."""
        output = str(tmp_path / "test_output.xlsx")
        with patch("parser.OUTPUT_FILE", output):
            save_to_excel(sample_progress)
            assert Path(output).exists()

    def test_correct_number_of_rows(self, tmp_path, sample_progress):
        """Количество строк = заголовок + количество университетов."""
        import openpyxl

        output = str(tmp_path / "test_output.xlsx")
        with patch("parser.OUTPUT_FILE", output):
            save_to_excel(sample_progress)

            wb = openpyxl.load_workbook(output)
            ws = wb.active
            # 1 заголовок + 2 университета = 3 строки
            assert ws.max_row == 3

    def test_headers_match_columns(self, tmp_path, sample_progress):
        """Заголовки в Excel совпадают с COLUMNS."""
        import openpyxl

        output = str(tmp_path / "test_output.xlsx")
        with patch("parser.OUTPUT_FILE", output):
            save_to_excel(sample_progress)

            wb = openpyxl.load_workbook(output)
            ws = wb.active

            headers = [ws.cell(row=1, column=i).value for i in range(1, len(COLUMNS) + 1)]
            assert headers == COLUMNS

    def test_data_values_correct(self, tmp_path, sample_progress):
        """Проверяем что числа из прогресса попали в правильные ячейки."""
        import openpyxl

        output = str(tmp_path / "test_output.xlsx")
        with patch("parser.OUTPUT_FILE", output):
            save_to_excel(sample_progress)

            wb = openpyxl.load_workbook(output)
            ws = wb.active

            # Строка 2 = Harvard
            assert ws.cell(row=2, column=1).value == "Harvard University"
            assert ws.cell(row=2, column=2).value == 36012  # Total students
            assert ws.cell(row=2, column=4).value == 7744   # International students


# ============================================================
# ТЕСТ 4: Параметризация — один тест, много входных данных
# ============================================================
# Это мощная фича pytest. Вместо копипасты 5 одинаковых тестов
# с разными данными — пишем один и даём ему список случаев.

@pytest.mark.parametrize("raw_label,expected_column", [
    ("Total number of students", "Total number of students"),
    ("Number of master&#x27;s degrees awarded", "Number of master's degrees awarded"),
    ("Number of new master&#x27;s students", "Number of new master's students"),
    ("Number of doctoral degrees awarded", "Number of doctoral degrees awarded"),
])
def test_stat_labels_mapping(raw_label, expected_column):
    """Проверяем что все HTML-метки маппятся на правильные колонки."""
    from html import unescape
    decoded = unescape(raw_label)
    assert decoded in STAT_LABELS or raw_label in STAT_LABELS


@pytest.mark.parametrize("value_str,expected_int", [
    ("36,012", 36012),
    ("7,744", 7744),
    ("1,234", 1234),
    ("500", 500),
    ("0", 0),
])
def test_number_cleaning(value_str, expected_int):
    """Проверяем что запятые в числах убираются корректно."""
    clean = value_str.replace(",", "")
    assert int(clean) == expected_int


# ============================================================
# ТЕСТ 5: Проверка констант (защита от случайных изменений)
# ============================================================

class TestConstants:

    def test_columns_count(self):
        """У нас ровно 13 колонок по шаблону заказчика."""
        assert len(COLUMNS) == 13

    def test_first_column_is_name(self):
        """Первая колонка — имя университета."""
        assert COLUMNS[0] == "University Name"

    def test_link_column_exists(self):
        """Колонка Link существует."""
        assert "Link" in COLUMNS

    def test_all_stat_labels_map_to_valid_columns(self):
        """Каждая метка из STAT_LABELS указывает на существующую колонку."""
        for label, column in STAT_LABELS.items():
            assert column in COLUMNS, f"'{column}' not in COLUMNS"
